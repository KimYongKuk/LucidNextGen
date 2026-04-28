"""XlsxWorker - Excel(XLSX) 파일 생성/수정 전담 Worker

담당 도구: excel-mcp-server의 24개 도구
- Workbook: create_workbook, create_worksheet, get_workbook_metadata
- Data: read_data_from_excel, write_data_to_excel
- Formula: apply_formula, validate_formula_syntax
- Format: format_range, merge_cells, unmerge_cells, get_merged_cells
- Chart: create_chart
- Pivot: create_pivot_table
- Table: create_table
- Sheet mgmt: copy_worksheet, delete_worksheet, rename_worksheet
- Row/Col: insert_rows, insert_columns, delete_sheet_rows, delete_sheet_columns
- Range: copy_range, delete_range, validate_excel_range, get_data_validation_info

Sonnet 모델 사용: 복잡한 자연어 → 다단계 Excel 도구 호출 변환 필요
"""

import asyncio
import copy
import re
import shutil
import time
import zipfile
import xml.etree.ElementTree as ET
from io import BytesIO
from pathlib import Path
from typing import List, Dict, Any, AsyncIterator, Optional

from langchain_core.messages import BaseMessage
from langchain_core.tools import BaseTool

from .base_worker import BaseWorker

# ============================================================================
# 경로 설정
# ============================================================================
_BACKEND_DATA = Path(__file__).parent.parent.parent.parent / "data"
XLSX_UPLOAD_DIR = _BACKEND_DATA / "xlsx_upload"
XLSX_OUTPUT_DIR = _BACKEND_DATA / "xlsx_output"

# 디렉토리 자동 생성
XLSX_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
XLSX_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# ============================================================================
# 파일별 asyncio Lock (LangGraph의 병렬 도구 실행으로 인한 동시 접근 방지)
# LangGraph는 한 LLM 응답에서 여러 tool_calls를 반환하면 ToolNode가 이를
# 병렬 실행 → 같은 xlsx 파일에 동시 load_workbook/save → ZIP 손상 발생
# ============================================================================
_file_locks: Dict[str, asyncio.Lock] = {}


# ============================================================================
# 세션별 작업 파일 추적 (디스크 영속화)
# ----------------------------------------------------------------------------
# "한 세션 내 연속 작업" 시나리오 (생성 → 수정 → 수정)를 지원하기 위해
# create_xlsx/modify_xlsx 성공 시 그 파일을 이 세션의 "작업 파일"로 등록.
# `_list_available_files`가 업로드 파일 + 이 세션 작업 파일만 노출하여
# 다른 세션의 output이 컨텍스트를 오염시키지 않도록 한다.
#
# 영속화: in-memory dict + JSON 인덱스 파일 (data/xlsx_session_index.json).
# 모듈 import 시 JSON 로드 → dict 복원. register 시 JSON atomic write.
# 백엔드 재시작 후에도 세션-파일 연결 유지.
# ============================================================================
_session_output_files: Dict[str, "set[str]"] = {}
_SESSION_INDEX_PATH = _BACKEND_DATA / "xlsx_session_index.json"


def _load_session_index() -> None:
    """모듈 import 시 디스크에서 dict 복원."""
    global _session_output_files
    try:
        if _SESSION_INDEX_PATH.exists():
            import json
            with open(_SESSION_INDEX_PATH, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, dict):
                _session_output_files = {
                    sid: set(paths) if isinstance(paths, list) else set()
                    for sid, paths in data.items()
                }
                print(f"[XlsxWorker] [SESSION_INDEX] 디스크에서 로드: {len(_session_output_files)} 세션")
    except Exception as e:
        print(f"[XlsxWorker] [SESSION_INDEX] 로드 실패 (무시하고 빈 상태로 시작): {e}")


def _flush_session_index() -> None:
    """dict를 디스크에 atomic write (temp + rename)."""
    try:
        import json
        _SESSION_INDEX_PATH.parent.mkdir(parents=True, exist_ok=True)
        serializable = {sid: sorted(paths) for sid, paths in _session_output_files.items()}
        tmp = _SESSION_INDEX_PATH.with_suffix(".json.tmp")
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(serializable, f, ensure_ascii=False, indent=2)
        tmp.replace(_SESSION_INDEX_PATH)
    except Exception as e:
        print(f"[XlsxWorker] [SESSION_INDEX] flush 실패 (메모리는 유지): {e}")


def _register_session_file(session_id: str, filepath: str) -> None:
    if not session_id or not filepath:
        return
    try:
        normalized = str(Path(filepath).resolve()).replace("\\", "/")
    except Exception:
        normalized = str(filepath).replace("\\", "/")
    bucket = _session_output_files.setdefault(session_id, set())
    if normalized in bucket:
        return  # no-op (디스크 쓰기 절약)
    bucket.add(normalized)
    _flush_session_index()


def _get_session_files(session_id: str) -> List[Path]:
    """이 세션에서 생성·수정된 파일 중 현재도 존재하는 것들을 최신순으로 반환."""
    if not session_id:
        return []
    paths = _session_output_files.get(session_id) or set()
    existing: List[Path] = []
    stale = []
    for p in paths:
        try:
            pth = Path(p)
            if pth.exists():
                existing.append(pth)
            else:
                stale.append(p)
        except Exception:
            stale.append(p)
    # 사라진 파일은 인덱스에서 정리 (기회적 GC)
    if stale:
        for s in stale:
            paths.discard(s)
        if not paths:
            _session_output_files.pop(session_id, None)
        _flush_session_index()
    existing.sort(key=lambda x: x.stat().st_mtime, reverse=True)
    return existing


# 모듈 import 시점에 인덱스 로드
_load_session_index()


def _get_file_lock(filepath: str) -> asyncio.Lock:
    """파일 경로별 asyncio.Lock 반환 (없으면 생성)"""
    # 경로 정규화하여 동일 파일에 대해 같은 Lock 보장
    normalized = str(Path(filepath).resolve()).replace("\\", "/").lower()
    if normalized not in _file_locks:
        _file_locks[normalized] = asyncio.Lock()
    return _file_locks[normalized]


# ============================================================================
# READ-ONLY 도구 목록 (파일을 수정하지 않는 도구)
# WRITE 도구가 upload 파일을 대상으로 할 때 output으로 자동 복사/리다이렉트
# ============================================================================
READ_ONLY_TOOLS = frozenset([
    "get_workbook_metadata",
    "read_data_from_excel",
    "get_merged_cells",
    "validate_excel_range",
    "get_data_validation_info",
    "validate_formula_syntax",
])

# Tool result 최대 길이 (개별 결과 안전망 — 극단적 대량 데이터 방어)
TOOL_RESULT_MAX_CHARS = 8000

# ============================================================================
# 쓰기 도구 성공 응답 표준화
# ----------------------------------------------------------------------------
# 배경: excel-mcp-server는 성공 응답이 극히 짧음
#   - create_workbook: "Created workbook at {path}"
#   - write_data_to_excel: "Data written to {sheet_name}"  (20자 내외)
# 짧고 모호한 응답이 ReAct 압축 및 GUARD 차단과 결합될 때, Sonnet이
# "뭔가 잘못됐다"고 오인하여 AttributeError 등을 환각하는 사례 발생.
# 모든 쓰기 성공 응답을 `✅ SUCCESS:` 고정 포맷으로 정규화하여 LLM이
# 성공/실패를 모호함 없이 해석하도록 강제한다.
# ============================================================================
_ERROR_PREFIXES = ("Error:", "❌", "Failed:", "ValueError", "WorkbookError", "DataError", "ValidationError")


def _is_error_response(text: str) -> bool:
    """MCP 도구 응답이 명시적 에러인지 판별 (접두사 기반)"""
    if not isinstance(text, str):
        return False
    stripped = text.lstrip()
    return any(stripped.startswith(p) for p in _ERROR_PREFIXES)


def _enrich_tool_result(tool_name: str, target_args: dict, result):
    """Excel 쓰기 도구 성공 응답을 `✅ SUCCESS:` 표준 포맷으로 정규화.

    - 에러 응답은 그대로 통과 (LLM이 재시도/안내 판단)
    - 읽기 전용 도구는 그대로 통과 (데이터 내용이 중요)
    - 쓰기 성공 응답에 파일명/수치/다음 단계 정보 주입
    """
    if not isinstance(result, str) or _is_error_response(result):
        return result
    if tool_name in READ_ONLY_TOOLS:
        return result

    # 합성 도구(xlsx_simple/server.py)는 이미 완전한 ✅ SUCCESS 포맷으로 반환하므로
    # 추가 enrich 불필요. 그대로 통과시켜 중복 SUCCESS 접두사 방지.
    if tool_name in ("create_xlsx", "modify_xlsx"):
        return result

    filepath = target_args.get("filepath", "") or ""
    filename = Path(filepath).name if filepath else "(파일명 미확인)"
    sheet_name = target_args.get("sheet_name", "") or ""

    if tool_name == "create_workbook":
        return (
            f"✅ SUCCESS: {result}\n"
            f"- 파일명: {filename}\n"
            f"- 기본 시트 'Sheet' 자동 생성됨 (빈 상태)\n"
            f"NEXT STEP: write_data_to_excel(filepath='{filepath}', sheet_name='Sheet', data=[[헤더...],[행1...],...])를 반드시 호출하세요. "
            f"create_workbook을 재호출하지 마세요."
        )

    if tool_name == "write_data_to_excel":
        data = target_args.get("data", [])
        rows = len(data) if isinstance(data, list) else 0
        cols = len(data[0]) if rows > 0 and isinstance(data[0], list) else 0
        return (
            f"✅ SUCCESS: {result}\n"
            f"- 파일: {filename}\n"
            f"- 시트: '{sheet_name}'\n"
            f"- 작성 범위: {rows}행 × {cols}열\n"
            f"작업이 정상 완료되었습니다. 사용자에게 `**파일명:** {filename}` 형태로 안내하세요."
        )

    if tool_name == "apply_formula":
        cell = target_args.get("cell", "")
        formula = target_args.get("formula", "")
        return (
            f"✅ SUCCESS: {result}\n"
            f"- 파일: {filename}, 시트: '{sheet_name}', 셀: {cell}\n"
            f"- 수식: {formula[:100]}"
        )

    if tool_name in ("create_worksheet", "rename_worksheet", "copy_worksheet", "delete_worksheet"):
        return f"✅ SUCCESS: {result}\n- 파일: {filename}"

    if tool_name in ("merge_cells", "unmerge_cells", "format_range",
                     "insert_rows", "insert_columns", "delete_sheet_rows", "delete_sheet_columns",
                     "copy_range", "delete_range"):
        return f"✅ SUCCESS: {result}\n- 파일: {filename}, 시트: '{sheet_name}'"

    if tool_name in ("create_chart", "create_pivot_table", "create_table"):
        return f"✅ SUCCESS: {result}\n- 파일: {filename}, 시트: '{sheet_name}'"

    # 기타 쓰기 도구 — 접두사만 추가
    return f"✅ SUCCESS: {result}"


class XlsxWorker(BaseWorker):
    """
    Excel(XLSX) 파일 생성/수정 Worker (Sonnet)

    Sonnet 사용 이유: 복잡한 자연어 → 다단계 Excel 도구 호출 변환 필요
    """

    @property
    def name(self) -> str:
        return "XlsxWorker"

    @property
    def tool_names(self) -> List[str]:
        # 핵심 원칙: xlsx 쓰기 도구는 합성 도구 2개(create_xlsx/modify_xlsx)뿐.
        # excel-mcp의 2-step 도구(create_workbook/write_data_to_excel 등)를 노출하지 않아
        # Sonnet의 multi-call 불안정성(중복 호출/경로 변조/환각) 여지 원천 차단.
        return [
            # 신규 엑셀 생성 — 단일 호출 완결 (단일 or 다중 시트)
            "create_xlsx",
            # 기존 엑셀 수정 — 단일 호출 완결 (operations 배열로 여러 변경 일괄 적용)
            "modify_xlsx",
            # 기존 파일 조회 (읽기 전용)
            "get_workbook_metadata",
            "read_data_from_excel",
            # 웹 검색 (시장 데이터·통계 등 최신 정보 필요 시)
            "tavily_search",
        ]

    @property
    def use_sonnet(self) -> bool:
        return True

    @property
    def max_agent_steps(self) -> int:
        """Excel 작업은 다단계 도구 호출 필요 (기본 20 → 30 = 최대 15회 도구 호출)"""
        return 30

    @property
    def compact_previous_results(self) -> bool:
        """이전 단계 Tool 결과 압축 활성화 (토큰 누적 방지)"""
        return True

    @property
    def skip_auto_file_tools(self) -> bool:
        """XlsxWorker는 get_workbook_metadata/read_data_from_excel로 충분.
        chunk 검색(search_user_files)이 추가되면 Sonnet이 재검증 루프에 빠져 환각 유발."""
        return True

    @property
    def summarization_prompt(self) -> str:
        return """다음 대화 내용을 Excel 파일 생성/수정을 위해 요약해줘.

## 요약 지침
1. 핵심 데이터, 숫자, 통계는 정확히 보존
2. 주요 주제와 요청 사항 포함
3. 테이블 데이터가 있으면 구조 유지 (헤더, 행, 열)
4. 사용자의 최종 요청 명확히 기록
5. 마크다운 형식으로 정리
6. 최대 800단어

## ⚠️ 중요 - Excel 관련 정보 보존:
- 사용자가 요청한 엑셀 구조 (시트명, 열 구성, 데이터 타입)
- 서식 요청 (셀 색상, 글꼴, 테두리 등)
- 이전에 생성/수정된 파일명이 있다면 기록
- 수식/피벗테이블 요청 사항

## 대화 내용:
{conversation}

---
## 요약:"""

    @property
    def system_prompt(self) -> str:
        return self._base_prompt

    @property
    def _base_prompt(self) -> str:
        return """You are an Excel specialist. 사용자 요청에 따라 엑셀 파일을 생성/수정합니다.

## ⛔ 절대 규칙: 모든 요청에 반드시 도구를 호출하라
도구 호출 없이 "적용했습니다" / "수정했습니다" / "변경했습니다"라고 응답하면 실제로 아무 변경도 안 됩니다.
첫 번째 응답에서 바로 도구를 호출하세요. 사전 안내("만들겠습니다") 금지.

## ⭐ 새 엑셀 파일 생성: **create_xlsx 단 하나만 호출**
단일 시트:
```
create_xlsx(filepath="파일명.xlsx", headers=["A","B","C","D"], rows=[[1,2,3,4], [5,6,7,8]])
```
다중 시트 (sheets 배열 사용):
```
create_xlsx(filepath="파일명.xlsx", sheets=[
  {"name":"매출", "headers":["월","금액"], "rows":[["1월",100],["2월",200]]},
  {"name":"비용", "headers":["항목","금액"], "rows":[["인건비",50]]},
])
```
- 이 도구 하나로 파일 생성 + 시트(들) + 데이터 쓰기 + 저장이 **한 번에** 완료됩니다.

## ⭐ 기존 엑셀 파일 수정: **modify_xlsx 단 하나만 호출**
```
modify_xlsx(filepath="기존파일.xlsx", operations=[...])
```
지원 op 15종 (한 번에 여러 개 섞어서 operations 배열에 넣으세요):

**[데이터]**
- `{"op":"update_cells", "sheet":"Sheet", "start_cell":"A2", "values":[[100,200]]}`
- `{"op":"apply_formula", "sheet":"Sheet", "cell":"C10", "formula":"=SUM(C2:C9)"}`

**[시트]**
- `{"op":"add_sheet", "name":"Summary", "headers":[...], "rows":[[...]]}`
- `{"op":"delete_sheet", "name":"Sheet2"}`
- `{"op":"rename_sheet", "old_name":"Sheet1", "new_name":"메인"}`
- `{"op":"copy_worksheet", "source":"Sheet1", "target":"Sheet1_복사본"}` — **서식·수식·병합 모두 포함** 시트 통째 복사

**[행·열]**
- `{"op":"insert_rows", "sheet":"Sheet", "start_row":2, "count":1}`
- `{"op":"insert_columns", "sheet":"Sheet", "start_col":"B", "count":1}` — 문자 또는 정수
- `{"op":"delete_rows", "sheet":"Sheet", "start_row":2, "count":3}`
- `{"op":"delete_columns", "sheet":"Sheet", "start_col":"B", "count":1}`

**[서식]**
- `{"op":"format_range", "sheet":"Sheet", "start_cell":"A1", "end_cell":"D1", "bold":true, "bg_color":"#FFFF00", "font_color":"#000000", "alignment":"center", "border_style":"thin", "number_format":"#,##0"}` — 모든 스타일 옵션 선택적
- `{"op":"merge_cells", "sheet":"Sheet", "start_cell":"A1", "end_cell":"D1"}`
- `{"op":"unmerge_cells", "sheet":"Sheet", "start_cell":"A1", "end_cell":"D1"}`

**[차트·피벗]**
- `{"op":"create_chart", "sheet":"Sheet", "chart_type":"bar"|"line"|"pie", "data_range":"A1:B10", "target_cell":"D2", "title":"월별 매출", "x_axis":"월", "y_axis":"금액"}`
- `{"op":"create_pivot_table", "source_sheet":"Data", "source_range":"A1:D100", "target_sheet":"Pivot", "rows":["지역"], "values":[["매출","sum"]], "agg_func":"sum"}`

- 여러 op를 한 배열에 넣으면 **한 번의 호출로 원자적 적용** (중간 실패 시 파일 원본 유지).
- "기존 시트 그대로 복사" 요청은 `copy_worksheet`를 쓰세요 (값만 복사하는 add_sheet와 달리 서식·수식까지 보존).

## ⚠️ 절대 규칙 — metadata 성공 = 파일 접근 성공
- `get_workbook_metadata`가 `{'filename': ..., 'sheets': [...]}`를 반환했다면 **파일 접근은 성공한 것**입니다.
- 이전 대화 히스토리에 "파일 접근 오류", "업로드 경로 접근 불가" 같은 문구가 있더라도 **그것은 과거 실패이며 현재 턴과 무관**합니다. metadata 응답의 시트명을 그대로 **신뢰**하고 즉시 `modify_xlsx`를 호출하세요.
- **`get_workbook_metadata`는 요청당 최대 1회만** 호출하세요. 같은 파일에 대해 여러 번 호출하지 마세요.
- "접근 오류", "서버 오류" 같은 추측성 응답 절대 금지. metadata가 성공했으면 파일은 접근 가능한 것입니다.

## ⚠️ 수정 대상 파일 규칙 — 업로드 파일만 사용
- 수정 요청 시 **대상은 위 `{available_files}` 목록의 업로드된 파일**입니다.
- 대화 히스토리나 user memory에 이전 세션의 결과물 경로(예: `xlsx_output/...복사본.xlsx`)가 있어도 **무시**하세요. 그것들은 현재 요청과 무관합니다.
- `get_workbook_metadata`는 업로드 파일 경로 하나만 쓰세요. 여러 후보를 metadata로 확인하지 마세요 — 그러면 혼란만 가중됩니다.

## ⚠️ 시트명 매칭 규칙 — fuzzy 매칭으로 진행
사용자가 말하는 시트명과 metadata로 받은 실제 시트명이 정확히 일치하지 않는 경우가 흔합니다 (예: 사용자 "중지된 사용자" vs 실제 "중지 리스트"). 이때:
- **의미상 가장 유사한 시트를 선택해서 즉시 진행**하세요. "중지된 사용자" → "중지 리스트", "매출표" → "매출", "직원" → "직원목록" 등.
- 한국어 동의어/축약/유사 키워드 매칭 적극 활용. metadata의 시트 목록 중 어느 하나가 사용자 의도와 합리적으로 매칭되면 그것을 사용.
- 응답에 "정확히는 '실제시트명'이라는 시트로 작업했습니다" 정도만 짧게 안내하면 됩니다.
- **"그런 시트 없음" / "파일 접근 오류" 같은 환각 절대 금지**. 시트가 metadata에 보였으면 접근은 성공한 것입니다.
- 정말 매칭 후보가 0개일 때만 metadata 시트 목록을 제시하며 사용자 확인 요청.

## 결정 플로우 (중요 — 순서대로 판정)
1. **새 파일 요청** → `create_xlsx` (1번)
2. **대화에 이전 테이블 데이터가 있는 수정 요청** → `create_xlsx`로 새로 생성 (덮어쓰기). 읽기 단계 불필요.
3. **시트 복사/백업/복제** (예: "시트 그대로 복사", "시트 복제", "같은 내용 시트 하나 더") →
   (a) `get_workbook_metadata`로 원본 시트명 확인 → (b) **바로** `modify_xlsx(operations=[{"op":"copy_worksheet", "source":..., "target":...}])`.
   **`read_data_from_excel` 호출 금지** — `copy_worksheet`가 내부에서 값·서식·수식·병합까지 전부 복사합니다.
4. **시트 관리 전용** (추가/삭제/이름변경, 행/열 삽입·삭제, 서식, 수식, 병합, 차트, 피벗) → 데이터 내용을 참조하지 않는 경우:
   (a) `get_workbook_metadata`로 시트명 확인 → (b) 바로 `modify_xlsx`. read_data_from_excel 스킵.
5. **기존 값을 참조해야 하는 수정** (예: "C열 합계 수식", "현재 값의 10% 추가") →
   (a) `get_workbook_metadata` → (b) `read_data_from_excel` → (c) `modify_xlsx`

## 파일
{available_files}

## 웹 검색 (tavily_search)
- 시장 데이터·통계·트렌드 등 최신 정보가 필요한 엑셀 생성 요청 시 먼저 조사. 직접 제공 데이터·기존 파일 수정은 불필요.

## 규칙
1. **경로**: 새 파일은 `{output_dir}/파일명.xlsx`. 기존 파일은 AVAILABLE FILES의 경로 사용.
2. **create_xlsx 또는 modify_xlsx 호출은 딱 1번**. `✅ SUCCESS:` 수신 즉시 파일명 안내하고 종료.
3. **data 형식**: rows는 `List[List]` 형태. 예: `[[1,2,3,4], [5,6,7,8]]`

## 응답 형식
- 한국어 응답. JSON/data 배열 노출 금지
- `✅ SUCCESS:` 수신 시: 간략 설명 + "**파일명:** xxx.xlsx"
- 에러 접두사(`Error:`, `❌`)가 없으면 성공. "서버 오류" 같은 추측 응답 금지.

Answer in Korean unless asked otherwise."""

    def build_system_prompt(
        self,
        context: Dict[str, Any],
        memory_context: Optional[Dict[str, Any]] = None,
        user_memory_context: Optional[Dict[str, Any]] = None,
    ) -> str:
        """컨텍스트를 반영한 시스템 프롬프트 생성"""
        prompt = self._base_prompt

        # output_dir 주입
        output_dir = str(XLSX_OUTPUT_DIR).replace("\\", "/")
        prompt = prompt.replace("{output_dir}", output_dir)

        # 세션의 업로드된 xlsx 파일 목록 탐색
        session_id = context.get("session_id", "")
        available_files_text = self._list_available_files(session_id)
        prompt = prompt.replace("{available_files}", available_files_text)

        # 세션 ID / 워크스페이스 UUID 주입
        if session_id:
            prompt = prompt.replace("{session_id}", session_id)
        workspace_uuid = context.get("workspace_uuid", "")
        if workspace_uuid:
            prompt = prompt.replace("{workspace_uuid}", workspace_uuid)

        # 날짜 정보
        from datetime import datetime
        now = datetime.now()
        weekdays = ["월요일", "화요일", "수요일", "목요일", "금요일", "토요일", "일요일"]
        weekday_kr = weekdays[now.weekday()]
        current_date = f"{now.year}년 {now.month}월 {now.day}일 ({weekday_kr})"
        prompt = f"Today is {current_date}.\n\n{prompt}"

        # 워크스페이스 instructions
        workspace_instructions = context.get("workspace_instructions")
        if workspace_instructions:
            prompt = f"WORKSPACE INSTRUCTIONS:\n{workspace_instructions}\n\n{prompt}"

        # 메모리 컨텍스트
        if memory_context and memory_context.get("summary"):
            summary = memory_context["summary"]
            key_facts = memory_context.get("key_facts", [])
            facts_text = "\n".join(f"  - {fact}" for fact in key_facts) if key_facts else "  (없음)"
            memory_section = f"\n## 워크스페이스 메모리\n이전 대화 요약: {summary}\n핵심 사실:\n{facts_text}\n"
            prompt = prompt + memory_section

        # 전역 사용자 메모리 주입
        if user_memory_context and user_memory_context.get("key_facts"):
            facts = user_memory_context["key_facts"]
            facts_text = "\n".join(f"  - {fact}" for fact in facts)
            prompt = f"## User Profile (사용자 개인 특성)\n\n이 사용자에 대해 알려진 정보:\n{facts_text}\n\n{prompt}"

        # ============ 공유 도구 안내 (PDF/DOCX/차트) ============
        prompt += """

## 다른 형식 문서 생성 (공유 도구)

사용자가 엑셀이 아닌 다른 형식의 문서를 요청할 수 있습니다.
- "워드로 만들어줘", "Word", "DOCX", "편집 가능한 문서" → `create_document_docx` 도구 호출
- "PDF로 만들어줘", "PDF로 정리해줘" → `create_document_pdf` 도구 호출
- 차트/그래프 요청 → `create_line_chart`, `create_bar_chart`, `create_pie_chart`, `create_multi_chart`

**핵심 규칙:**
1. "도구가 없습니다", "Excel만 가능합니다", "제 영역 밖입니다"라고 **절대 답하지 마세요**. 위 도구를 사용할 수 있습니다.
2. 도구 호출 시 content는 마크다운 형식으로 작성하세요.
3. 표 데이터는 마크다운 테이블(| col1 | col2 |) 형식으로 포함하세요.
"""

        print(f"[XlsxWorker] Context: session_id={bool(session_id)}, workspace_uuid={bool(workspace_uuid)}")

        return prompt

    def _list_available_files(self, session_id: str) -> str:
        """이 세션의 엑셀 파일 목록을 **업로드 원본**과 **이번 세션 작업 파일**로
        구분하여 반환. 다른 세션의 output은 노출하지 않음 (컨텍스트 오염 방지).

        - 업로드 원본: `xlsx_upload/{session_id}/` 하위의 모든 xlsx
        - 이번 세션 작업 파일: `_session_output_files[session_id]`에 등록된 파일들
          (create_xlsx/modify_xlsx 성공 시 secured_ainvoke가 등록)
        """
        uploaded: List[str] = []
        if session_id:
            upload_dir = XLSX_UPLOAD_DIR / session_id
            if upload_dir.exists():
                for f in upload_dir.glob("*.xlsx"):
                    uploaded.append(str(f).replace("\\", "/"))
                for f in upload_dir.glob("*.xls"):
                    uploaded.append(str(f).replace("\\", "/"))

        session_work = _get_session_files(session_id)

        if not uploaded and not session_work:
            return "(이번 세션에 업로드/작업된 엑셀 파일이 없습니다. 새 파일은 `create_xlsx`로 생성하세요.)"

        lines: List[str] = []
        if uploaded:
            lines.append("**업로드 원본** (읽기 및 수정 가능):")
            for p in uploaded:
                lines.append(f"- {p}")
        if session_work:
            lines.append("")
            lines.append("**이번 세션 작업 파일** (create_xlsx/modify_xlsx로 만들거나 수정한 결과, 최신순):")
            for p in session_work:
                lines.append(f"- {str(p).replace(chr(92), '/')}")
            lines.append("")
            lines.append("→ 연속 편집 시 **가장 최근 작업 파일**을 filepath로 사용하세요. 이전 세션의 output은 노출되지 않습니다.")

        return "\n".join(lines)

    # ==========================================================
    # prepare_tools: filepath 보안 래핑 (핵심!)
    # ==========================================================

    def prepare_tools(
        self,
        tools: List[BaseTool],
        context: Dict[str, Any],
    ) -> List[BaseTool]:
        """
        모든 Excel 도구의 filepath 파라미터를 검증/샌드박싱

        허용 디렉토리:
        1. XLSX_UPLOAD_DIR/{session_id}/ (업로드된 원본 파일)
        2. XLSX_OUTPUT_DIR/ (생성/수정된 파일)

        보안:
        - ".." 경로 탐색 차단
        - 허용 디렉토리 외 접근 차단
        - 상대 경로를 XLSX_OUTPUT_DIR 기준으로 자동 해석
        """
        session_id = context.get("session_id", "")
        user_id = context.get("user_id", "") or "unknown"
        allowed_dirs = [str(XLSX_OUTPUT_DIR.resolve())]
        if session_id:
            upload_dir = XLSX_UPLOAD_DIR / session_id
            upload_dir.mkdir(parents=True, exist_ok=True)
            allowed_dirs.append(str(upload_dir.resolve()))

        # Per-request: upload 파일이 output으로 복사되면 이후 모든 도구가 output 경로 사용
        redirected_files: Dict[str, str] = {}

        # create_workbook 중복 호출 차단용 (단일 파일 생성 후 재호출 시 GUARD)
        # 신규 생성은 create_xlsx 합성 도구가 담당하므로 이 경로는 점점 덜 쓰임.
        created_workbook: Dict[str, Optional[str]] = {"path": None}

        # Circuit breaker: 파일 생성 성공 후 추가 xlsx 도구 호출 전부 차단.
        # Sonnet 4.6이 성공 응답을 받고도 의심하여 재호출/다른 도구 시도하는 behavior 방어.
        # 한 번 "xlsx 생성 완료" 상태가 되면, 이후 쓰기 도구 호출은 전부 동일한
        # 확정 메시지로 short-circuit하여 LLM이 최종 text 응답을 생성하도록 강제.
        creation_done: Dict[str, Any] = {"file": None}
        XLSX_WRITE_TOOLS = frozenset([
            "create_xlsx",
            "modify_xlsx",
            "create_workbook",
            "write_data_to_excel",
            "apply_formula",
            "format_range",
            "merge_cells",
            "unmerge_cells",
            "create_worksheet",
            "rename_worksheet",
            "copy_worksheet",
            "delete_worksheet",
            "create_chart",
            "create_pivot_table",
            "create_table",
            "insert_rows",
            "insert_columns",
            "delete_sheet_rows",
            "delete_sheet_columns",
            "copy_range",
            "delete_range",
        ])

        # 보안 래핑 대상: Excel 전용 도구만 (tavily_search 등 외부 도구는 제외)
        # MCP 도구는 전역 캐시되므로, 외부 도구를 래핑하면 다른 Worker에도 영향
        SKIP_WRAPPING = frozenset(["tavily_search"])

        # MCP 도구는 글로벌 캐시되어 모든 요청이 공유한다. 직접 ainvoke 덮어쓰기는
        # 동시 요청 race로 다른 사용자/세션의 redirected_files·created_workbook·creation_done
        # closure가 섞이는 누설을 일으킨다. → 사용자별 사본을 만들고 사본의 ainvoke만 교체.
        prepared: List[BaseTool] = []
        for tool in tools:
            if tool.name in SKIP_WRAPPING:
                # 이전 패치 잔재 정리: 글로벌 객체에 박힌 wrapper가 있으면 원복
                unwrapped = getattr(tool, "_unwrapped_ainvoke", None)
                if unwrapped:
                    object.__setattr__(tool, "ainvoke", unwrapped)
                    print(f"[XlsxWorker] [UNWRAP] {tool.name}: 이전 래핑 해제됨")
                prepared.append(tool)
                continue

            user_tool = copy.copy(tool)
            original_ainvoke = tool.ainvoke

            async def secured_ainvoke(
                input_data,
                config=None,
                *,
                _original=original_ainvoke,
                _allowed=allowed_dirs,
                _output_dir=str(XLSX_OUTPUT_DIR),
                _tname=tool.name,
                _redirected=redirected_files,
                _user_id_for_archive=user_id,
                _session_id_for_track=session_id,
                _created=created_workbook,
                _done=creation_done,
                _write_tools=XLSX_WRITE_TOOLS,
                **kwargs,
            ):
                # Circuit breaker: 이미 xlsx 파일 생성이 완료된 상태에서 xlsx 쓰기 도구가
                # 또 호출되면, 실제 실행 없이 확정 성공 메시지로 short-circuit.
                # LLM이 성공을 의심하여 재호출/다른 도구 시도하는 behavior를 코드로 차단.
                if _done.get("file") and _tname in _write_tools:
                    done_file = _done["file"]
                    print(f"[XlsxWorker] [CIRCUIT_BREAKER] {_tname} 호출 무효화 → 이미 생성됨: {Path(done_file).name}")
                    return (
                        f"✅ SUCCESS: 파일 생성이 이미 완료되었습니다.\n"
                        f"- 파일명: {Path(done_file).name}\n"
                        f"STOP: 추가 도구 호출이 필요하지 않습니다. "
                        f"즉시 사용자에게 `**파일명:** {Path(done_file).name}` 형태로 안내하고 응답을 종료하세요."
                    )

                resolved_filepath = None

                if isinstance(input_data, dict):
                    # ToolCall format: {name, args, id, type}
                    target = input_data.get("args", input_data) if "args" in input_data else input_data

                    # 디버깅: write_data_to_excel의 data 형식 확인
                    if _tname == "write_data_to_excel" and "data" in target:
                        data_val = target["data"]
                        if isinstance(data_val, list) and len(data_val) > 0:
                            first_item = data_val[0]
                            print(f"[XlsxWorker] [DEBUG] write_data data type: List[{type(first_item).__name__}], rows={len(data_val)}, first_row={first_item}")
                        else:
                            print(f"[XlsxWorker] [DEBUG] write_data data: {type(data_val).__name__}, value={str(data_val)[:200]}")

                    if "filepath" in target:
                        raw_path = target["filepath"]
                        validated = _validate_filepath(raw_path, _allowed, _output_dir)

                        # upload 파일 수정 시 output으로 자동 복사/리다이렉트
                        validated = _redirect_upload_to_output(
                            validated, _tname, _redirected, _output_dir
                        )

                        # create_workbook 중복 호출 차단 + 덮어쓰기 전 archive 백업
                        if _tname == "create_workbook":
                            if _created.get("path"):
                                prev = _created["path"]
                                raw_msg = f"Created workbook at {prev}"
                                print(f"[XlsxWorker] [GUARD] create_workbook 중복 호출 → 기존 파일 재확인: {Path(prev).name}")
                                return _enrich_tool_result(
                                    "create_workbook",
                                    {"filepath": prev},
                                    raw_msg,
                                )
                            _archive_previous_version(validated, _user_id_for_archive)
                            _created["path"] = validated

                        target["filepath"] = validated
                        resolved_filepath = validated
                        print(f"[XlsxWorker] [SECURE] {_tname}: filepath '{raw_path}' -> '{validated}'")

                # 파일별 Lock으로 동시 접근 직렬화
                # (LangGraph ToolNode가 병렬 실행해도 같은 파일은 순차 처리)
                try:
                    if resolved_filepath:
                        lock = _get_file_lock(resolved_filepath)
                        async with lock:
                            print(f"[XlsxWorker] [LOCK] {_tname}: acquired lock for {Path(resolved_filepath).name}")
                            result = await _original(input_data, config, **kwargs)
                            # create_workbook 성공 시 기본 시트명을 'Sheet'로 통일
                            # (excel-mcp는 'Sheet1'로 만들지만 프롬프트는 'Sheet' 사용)
                            if (
                                _tname == "create_workbook"
                                and isinstance(result, str)
                                and not _is_error_response(result)
                            ):
                                _normalize_default_sheet_name(resolved_filepath)
                            print(f"[XlsxWorker] [LOCK] {_tname}: released lock for {Path(resolved_filepath).name}")
                    else:
                        result = await _original(input_data, config, **kwargs)
                except Exception as e:
                    print(f"[XlsxWorker] [ERROR] {_tname}: {type(e).__name__}: {e}")
                    raise

                # 모든 쓰기 도구 성공 응답을 `✅ SUCCESS:` 표준 포맷으로 정규화
                # (근본 문제: excel-mcp의 터스한 응답이 LLM 환각 유발 → 구조적 방어)
                if isinstance(input_data, dict):
                    target_args = input_data.get("args", input_data) if "args" in input_data else input_data
                    if isinstance(target_args, dict):
                        result = _enrich_tool_result(_tname, target_args, result)

                # Circuit breaker 플래그 설정: create_xlsx 또는 write_data_to_excel이 성공하면
                # 이후 xlsx 도구 호출을 전부 short-circuit 대상으로 만듦.
                if (
                    _tname in ("create_xlsx", "write_data_to_excel")
                    and isinstance(result, str)
                    and not _is_error_response(result)
                    and resolved_filepath
                ):
                    _done["file"] = resolved_filepath
                    print(f"[XlsxWorker] [DONE] xlsx 생성 완료 플래그 설정 → {Path(resolved_filepath).name}")

                # 세션 작업 파일 추적: create_xlsx/modify_xlsx 성공 시 이 세션의 파일로 등록
                # 다음 턴의 `_list_available_files`에 노출되어 연속 편집을 지원.
                if (
                    _tname in ("create_xlsx", "modify_xlsx")
                    and isinstance(result, str)
                    and not _is_error_response(result)
                    and resolved_filepath
                ):
                    _register_session_file(_session_id_for_track, resolved_filepath)
                    print(f"[XlsxWorker] [SESSION_FILE] {_tname} → 세션 파일 등록: {Path(resolved_filepath).name} (session={_session_id_for_track[:8] if _session_id_for_track else 'none'})")

                # 긴 도구 결과 잘라서 토큰 폭증 방지 (Approach A)
                return _truncate_tool_result(result, _tname)

            object.__setattr__(user_tool, "ainvoke", secured_ainvoke)
            prepared.append(user_tool)

        print(f"[XlsxWorker] 보안 래핑 완료: {len(prepared)}개 도구, allowed_dirs={allowed_dirs}")

        # 아카이브 래핑 (보안 래핑 위에 적용 — Output 파일 복사)
        prepared = self._wrap_tools_for_archive(prepared, context)

        return prepared

    # ==========================================================
    # stream_response: BaseWorker 기본 + 후처리 (빈 시트 제거, 수식 캐시)
    # Haiku 요약은 BaseWorker.stream_response()에서 자동 처리
    # ==========================================================

    async def stream_response(
        self,
        messages: List[BaseMessage],
        context: Dict[str, Any],
        all_tools: List[BaseTool],
        memory_context: Optional[Dict[str, Any]] = None,
        user_memory_context: Optional[Dict[str, Any]] = None,
    ) -> AsyncIterator[Dict[str, Any]]:
        """
        BaseWorker 스트리밍 + 최종 방어.

        create_xlsx 또는 modify_xlsx 성공 감지 시, 이후 LLM의 text 응답을 무시하고
        결정론 메시지로 교체. Sonnet 4.6이 tool 성공 후에도 "서버 오류" 환각을
        뱉는 model-level 문제 대응.
        """
        from langchain_core.messages import AIMessageChunk

        success_filename: Optional[str] = None
        success_tool: Optional[str] = None
        suppress_llm_text: bool = False
        replacement_sent: bool = False

        async for event in super().stream_response(
            messages, context, all_tools, memory_context, user_memory_context
        ):
            event_kind = event.get("event", "")

            # 합성 도구 성공 감지 (tool_end 이벤트)
            if event_kind == "on_tool_end" and not success_filename:
                tool_name = event.get("name", "") or event.get("data", {}).get("name", "")
                if tool_name in ("create_xlsx", "modify_xlsx"):
                    output = event.get("data", {}).get("output")
                    text = str(output.content) if hasattr(output, "content") else str(output or "")
                    if "✅ SUCCESS" in text:
                        # 파일명은 공백 포함 가능하므로 .xlsx까지 lazy match
                        # 예: "- 파일명: (IT운영팀) 3월 LFON 계정생성 및 중지 이력_복사본.xlsx"
                        m = re.search(r"파일명[:：]\s*`?(.+?\.xlsx)`?(?:\s|$)", text)
                        if m:
                            success_filename = m.group(1).strip()
                            success_tool = tool_name
                            suppress_llm_text = True
                            print(f"[XlsxWorker] [FINAL_GUARD] {tool_name} 성공 감지 → LLM text 교체 모드: {success_filename}")

            # LLM text 억제: 성공 감지 후의 LLM 스트림/완료 이벤트는 버리고 한 번만 결정론 응답 주입
            if suppress_llm_text and event_kind in ("on_chat_model_stream", "on_chat_model_end"):
                if not replacement_sent and event_kind == "on_chat_model_stream":
                    replacement_sent = True
                    verb = "생성" if success_tool == "create_xlsx" else "수정"
                    replacement = (
                        f"엑셀 파일이 성공적으로 {verb}되었습니다.\n\n"
                        f"**파일명:** {success_filename}"
                    )
                    fake_event = dict(event)
                    fake_event["data"] = {"chunk": AIMessageChunk(content=replacement)}
                    yield fake_event
                continue

            yield event

        # Post-processing: remove empty sheets, then pre-compute formula values
        self._cleanup_empty_sheets()
        self._precompute_formulas()

    def _cleanup_empty_sheets(self) -> None:
        """Agent 완료 후: output 디렉토리의 최근 xlsx 파일에서 빈 시트 제거

        LLM이 create_workbook 후 기본 "Sheet"를 비워두고 새 시트에 데이터를 쓰는 경우,
        다운로드된 파일에 빈 시트가 남는 문제를 해결.
        """
        try:
            import openpyxl
            # 최근 60초 이내 수정된 파일만 처리
            cutoff = time.time() - 60
            for f in XLSX_OUTPUT_DIR.glob("*.xlsx"):
                if f.stat().st_mtime < cutoff:
                    continue
                try:
                    wb = openpyxl.load_workbook(str(f))
                    if len(wb.sheetnames) <= 1:
                        wb.close()
                        continue

                    empty_sheets = []
                    for name in wb.sheetnames:
                        ws = wb[name]
                        if ws.max_row <= 1 and ws.max_column <= 1 and ws.cell(1, 1).value is None:
                            empty_sheets.append(name)

                    # 전체가 빈 경우 건드리지 않음
                    if len(empty_sheets) < len(wb.sheetnames):
                        for name in empty_sheets:
                            del wb[name]
                            print(f"[XlsxWorker] [CLEANUP] Removed empty sheet '{name}' from {f.name}")
                        if empty_sheets:
                            wb.save(str(f))

                    wb.close()
                except Exception as e:
                    print(f"[XlsxWorker] [CLEANUP] Failed to process {f.name}: {e}")
        except Exception as e:
            print(f"[XlsxWorker] [CLEANUP] Error: {e}")

    def _precompute_formulas(self) -> None:
        """Agent 완료 후: output xlsx 파일의 수식 cached value를 주입

        openpyxl(excel-mcp-server)은 수식(<f>)은 저장하지만 계산값(<v>)은
        기록하지 않음. 프리뷰 뷰어(Univer/SheetJS)가 수식을 계산하지 못하므로,
        formulas 라이브러리로 계산 후 xlsx XML에 <v> 태그를 직접 주입한다.

        수식은 그대로 보존되어 다운로드한 파일에서 Excel로 열면 수식이 유지됨.
        """
        try:
            import formulas
            import numpy as np
        except ImportError:
            print("[XlsxWorker] [FORMULA] formulas/numpy not installed, skipping")
            return

        cutoff = time.time() - 60
        for f in XLSX_OUTPUT_DIR.glob("*.xlsx"):
            if f.stat().st_mtime < cutoff:
                continue
            try:
                # 1) formulas 라이브러리로 수식 계산
                xl_model = formulas.ExcelModel().loads(str(f)).finish()
                sol = xl_model.calculate()

                # 2) solution → {(SHEET_NAME, CELL_REF): scalar_value} 매핑
                computed: Dict[tuple, Any] = {}
                for key, ranges_obj in sol.items():
                    m = re.match(r"'\[.*?\](.*?)'!(.*)", str(key))
                    if not m:
                        continue
                    sheet_name = m.group(1).upper()
                    cell_ref = m.group(2).upper()
                    if ":" in cell_ref:  # range ref 스킵
                        continue

                    val = ranges_obj.value
                    if isinstance(val, np.ndarray):
                        val = val.flat[0] if val.size > 0 else None
                    elif isinstance(val, (list, tuple)):
                        try:
                            val = val[0][0]
                        except (IndexError, TypeError):
                            val = None

                    # numpy 타입 → Python native
                    if isinstance(val, (np.integer,)):
                        val = int(val)
                    elif isinstance(val, (np.floating,)):
                        val = float(val)

                    # 수식 결과가 아닌 일반 문자열은 스킵
                    if val is not None and not isinstance(val, str):
                        computed[(sheet_name, cell_ref)] = val

                if not computed:
                    continue

                # 3) xlsx XML에 <v> 태그 주입 (수식 보존)
                injected = _inject_cached_values(str(f), computed)
                if injected > 0:
                    print(f"[XlsxWorker] [FORMULA] Injected {injected} cached values into {f.name}")
            except Exception as e:
                print(f"[XlsxWorker] [FORMULA] Skipped {f.name}: {e}")



def _truncate_tool_result(
    result,
    tool_name: str,
    max_chars: int = TOOL_RESULT_MAX_CHARS,
):
    """개별 도구 결과가 max_chars 초과 시 잘라서 토큰 폭증 방지.

    모든 타입(str, list, dict, ToolMessage 등)을 문자열로 변환 후 잘라냄.
    LLM에게 잘렸음을 안내하여 정확성 유지.

    주의: Excel 도구 전용. 외부 도구(tavily_search 등)에는 적용하지 않음.
    """
    # 방어: Excel 도구가 아닌 경우 원본 반환 (전역 캐시 래핑 누수 방지)
    if tool_name == "tavily_search":
        return result
    # 문자열로 변환
    if isinstance(result, str):
        text = result
    elif hasattr(result, "content"):
        # ToolMessage 등 content 속성이 있는 객체
        text = result.content if isinstance(result.content, str) else str(result.content)
    else:
        text = str(result)

    if len(text) <= max_chars:
        # 원래 타입이 str이 아니었으면 변환된 텍스트를 반환
        return result if isinstance(result, str) else text

    original_len = len(text)
    truncated = text[:max_chars].rstrip()

    notice = (
        f"\n\n(참고: 응답이 길어 프리뷰는 {max_chars:,}자까지만 표시됩니다. 전체 데이터는 {original_len:,}자로 "
        f"파일에 정상적으로 존재합니다. 이는 오류가 아닙니다.)"
    )
    if tool_name == "read_data_from_excel":
        notice += (
            " modify_xlsx 도구는 이 제약 없이 파일 전체를 직접 처리하므로, 필요한 operations를 "
            "그대로 호출하세요."
        )

    print(f"[XlsxWorker] [TRUNCATE] {tool_name}: {original_len:,}자 → {max_chars:,}자로 잘림")
    return truncated + notice


def _normalize_default_sheet_name(filepath: str) -> bool:
    """`create_workbook` 직후 기본 시트 'Sheet1'을 'Sheet'로 rename.

    근본 원인 (2026-04-20 3차 회고):
      - `excel_mcp.workbook.create_workbook`의 sheet_name 기본값은 `"Sheet1"`
      - 프롬프트/LLM은 `sheet_name='Sheet'`를 사용 (기존 컨벤션)
      - write_data_to_excel이 'Sheet' 시트를 찾지 못해 **새로 생성** → 'Sheet1'(빈) + 'Sheet'(데이터) 공존
      - LLM이 get_workbook_metadata로 확인 시 시트 2개 보여서 "이상함" 감지 → 환각

    해결: create_workbook 직후 내부적으로 시트명을 'Sheet'로 통일.
    LLM은 시트명 변환을 몰라도 되고, write_data는 정상 동작.
    """
    try:
        import openpyxl
        p = Path(filepath)
        if not p.exists():
            return False
        wb = openpyxl.load_workbook(str(p))
        renamed = False
        if "Sheet1" in wb.sheetnames and "Sheet" not in wb.sheetnames:
            wb["Sheet1"].title = "Sheet"
            wb.save(str(p))
            renamed = True
            print(f"[XlsxWorker] [NORMALIZE_SHEET] '{p.name}': Sheet1 → Sheet")
        wb.close()
        return renamed
    except Exception as e:
        print(f"[XlsxWorker] [NORMALIZE_SHEET] 실패 (계속 진행): {e}")
        return False


def _archive_previous_version(filepath: str, user_id: str) -> None:
    """`create_workbook`이 기존 파일을 덮어쓰기 전에 이전 버전을 archive로 백업.

    이렇게 해야 DEDUP 제거 후에도 이전 파일이 복구 가능하다.
    archive_file은 날짜/사용자별 디렉토리에 copy하므로 덮어쓰기 발생 시에도
    같은 날 여러 번 생성된 파일은 마지막 버전만 archive에 남는 제약은 있음.
    (완벽한 version history 요구 시 별도 스키마 필요 — 현재 범위 밖)
    """
    try:
        p = Path(filepath)
        if not p.exists():
            return
        from app.utils.file_archive import archive_file
        archived = archive_file(str(p), user_id, file_type="xlsx")
        if archived:
            print(f"[XlsxWorker] [ARCHIVE_BEFORE_OVERWRITE] '{p.name}' → {archived}")
    except Exception as e:
        print(f"[XlsxWorker] [ARCHIVE_BEFORE_OVERWRITE] 백업 실패 (계속 진행): {e}")


def _deduplicate_filepath(filepath: str) -> str:
    """
    [DEPRECATED 2026-04-20] — DEDUP이 LLM mental model을 파괴하여 제거됨.
    함수 자체는 향후 다른 용도를 위해 유지.

    파일이 이미 존재하면 _1, _2 등 접미사를 붙여 중복 방지.

    예: 매출보고서.xlsx → 매출보고서_1.xlsx → 매출보고서_2.xlsx
    """
    p = Path(filepath)
    if not p.exists():
        return filepath

    stem = p.stem
    suffix = p.suffix
    parent = p.parent
    counter = 1
    while True:
        new_path = parent / f"{stem}_{counter}{suffix}"
        if not new_path.exists():
            print(f"[XlsxWorker] [DEDUP] '{p.name}' exists → renamed to '{new_path.name}'")
            return str(new_path).replace("\\", "/")
        counter += 1


def _validate_filepath(raw_path: str, allowed_dirs: list, default_dir: str) -> str:
    """
    filepath를 검증하고 안전한 절대 경로로 변환

    Rules:
    1. ".." 포함 시 거부
    2. 절대 경로면 allowed_dirs 안에 있는지 확인
    3. 상대 경로면 default_dir (xlsx_output) 기준으로 해석
    4. 파일명만 있으면 default_dir에 배치
    """
    # .. 차단
    if ".." in raw_path:
        raise ValueError(f"Path traversal detected: {raw_path}")

    # 정규화
    normalized = raw_path.replace("\\", "/")

    # 파일명만 있는 경우 (경로 구분자 없음)
    if "/" not in normalized:
        return str(Path(default_dir) / normalized).replace("\\", "/")

    # 절대 경로 체크
    p = Path(normalized)
    if p.is_absolute():
        resolved = str(p.resolve()).replace("\\", "/")
        for allowed in allowed_dirs:
            allowed_normalized = allowed.replace("\\", "/")
            if resolved.startswith(allowed_normalized):
                return normalized
        raise ValueError(f"Path outside allowed directories: {raw_path}")

    # 상대 경로 → default_dir 기준
    full_path = Path(default_dir) / normalized
    resolved = str(full_path.resolve()).replace("\\", "/")
    for allowed in allowed_dirs:
        allowed_normalized = allowed.replace("\\", "/")
        if resolved.startswith(allowed_normalized):
            return str(full_path).replace("\\", "/")

    raise ValueError(f"Path outside allowed directories: {raw_path}")


def _redirect_upload_to_output(
    validated_path: str,
    tool_name: str,
    redirected: Dict[str, str],
    output_dir: str,
) -> str:
    """
    Upload 디렉토리의 파일에 쓰기 작업 시 자동으로 output 디렉토리로 복사/리다이렉트.

    흐름:
    1. 이미 리다이렉트된 파일 → 캐싱된 output 경로 반환 (READ/WRITE 모두)
    2. upload 파일이 아님 → 그대로 반환
    3. READ-ONLY 도구 + 아직 리다이렉트 안 됨 → 그대로 반환
    4. WRITE 도구 + upload 파일 → output으로 복사 후 output 경로 반환

    이렇게 하면:
    - 원본 upload 파일이 보존됨
    - 수정된 파일이 xlsx_output/에 위치하여 다운로드 가능
    - 첫 write 이후 모든 read/write가 output 복사본을 대상으로 함
    """
    normalized = str(Path(validated_path).resolve()).replace("\\", "/").lower()

    # 1. 이미 리다이렉트된 파일이면 캐싱된 경로 반환
    if normalized in redirected:
        redirected_path = redirected[normalized]
        print(f"[XlsxWorker] [REDIRECT] {tool_name}: reusing output copy -> {Path(redirected_path).name}")
        return redirected_path

    # 2. upload 디렉토리 파일인지 확인
    upload_dir_normalized = str(XLSX_UPLOAD_DIR.resolve()).replace("\\", "/").lower()
    if not normalized.startswith(upload_dir_normalized):
        return validated_path  # upload 파일이 아님

    # 3. READ-ONLY 도구는 아직 복사 전이면 원본 그대로 읽기
    if tool_name in READ_ONLY_TOOLS:
        return validated_path

    # 4. WRITE 작업 → upload 파일을 output으로 복사
    src_path = Path(validated_path)
    if not src_path.exists():
        return validated_path  # 파일 없으면 복사 불가

    dst_path = Path(output_dir) / src_path.name
    dst_str = str(dst_path).replace("\\", "/")

    # output에 이미 같은 이름이 있으면 덮어쓰기 (이전 요청 결과물)
    try:
        shutil.copy2(str(src_path), dst_str)
        print(f"[XlsxWorker] [REDIRECT] Copied '{src_path.name}' to xlsx_output/ for modification")
    except Exception as e:
        print(f"[XlsxWorker] [REDIRECT] Copy failed: {e}, using original path")
        return validated_path

    # 리다이렉트 기록 (이후 모든 도구 호출에서 이 경로 사용)
    redirected[normalized] = dst_str
    return dst_str


# ============================================================================
# 수식 cached value 주입 (xlsx XML 직접 수정)
# ============================================================================

_SHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"


def _inject_cached_values(filepath: str, computed: Dict[tuple, Any]) -> int:
    """xlsx XML의 수식 셀에 <v>(cached value) 태그를 주입한다.

    수식(<f>)은 그대로 보존하면서 계산된 값(<v>)만 추가하여,
    프리뷰에서 값이 표시되고 다운로드 시 수식도 유지된다.

    Returns:
        주입된 셀 수
    """
    ET.register_namespace("", _SHEET_NS)

    with zipfile.ZipFile(filepath, "r") as zin:
        all_names = {item.filename for item in zin.infolist()}

        # workbook.xml.rels에서 rId → sheet XML 경로 매핑
        rels_xml = ET.fromstring(zin.read("xl/_rels/workbook.xml.rels"))
        rid_to_target: Dict[str, str] = {}
        for rel in rels_xml:
            target = rel.get("Target", "")
            # 절대(/xl/...) vs 상대(worksheets/...) 경로 정규화
            if target.startswith("/"):
                target = target[1:]
            elif not target.startswith("xl/"):
                target = "xl/" + target
            rid_to_target[rel.get("Id", "")] = target

        # workbook.xml에서 시트명 → rId 매핑
        wb_xml = ET.fromstring(zin.read("xl/workbook.xml"))
        sheets_el = wb_xml.find(f"{{{_SHEET_NS}}}sheets")
        if sheets_el is None:
            return 0

        sheet_xml_paths: Dict[str, str] = {}
        for s_el in sheets_el.findall(f"{{{_SHEET_NS}}}sheet"):
            rid = s_el.get(f"{{{_REL_NS}}}id", "")
            name = s_el.get("name", "").upper()
            target = rid_to_target.get(rid, "")
            if target in all_names:
                sheet_xml_paths[name] = target

        # 각 시트 XML에서 수식 셀의 <v> 태그 주입
        modified_files: Dict[str, bytes] = {}
        total_injected = 0

        for sheet_name_upper, xml_path in sheet_xml_paths.items():
            tree = ET.fromstring(zin.read(xml_path))
            sheet_data = tree.find(f"{{{_SHEET_NS}}}sheetData")
            if sheet_data is None:
                continue

            changed = False
            for row_el in sheet_data.findall(f"{{{_SHEET_NS}}}row"):
                for cell_el in row_el.findall(f"{{{_SHEET_NS}}}c"):
                    ref = cell_el.get("r", "").upper()
                    f_el = cell_el.find(f"{{{_SHEET_NS}}}f")
                    if f_el is None:
                        continue

                    key = (sheet_name_upper, ref)
                    if key not in computed:
                        continue

                    val = computed[key]
                    v_el = cell_el.find(f"{{{_SHEET_NS}}}v")
                    # <v>가 없거나 비어있으면 값 주입
                    if v_el is None:
                        v_el = ET.SubElement(cell_el, f"{{{_SHEET_NS}}}v")
                    if not v_el.text:
                        v_el.text = str(val)
                        changed = True
                        total_injected += 1

            if changed:
                modified_files[xml_path] = ET.tostring(
                    tree, xml_declaration=True, encoding="UTF-8"
                )

        if not modified_files:
            return 0

        # 수정된 시트만 교체하여 ZIP 재작성
        tmp_path = filepath + ".tmp"
        with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                if item.filename in modified_files:
                    zout.writestr(item, modified_files[item.filename])
                else:
                    zout.writestr(item, zin.read(item.filename))

    shutil.move(tmp_path, filepath)
    return total_injected
