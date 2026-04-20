"""
XLSX Simple MCP Server — single-call 합성 도구로 엑셀 전체 유스케이스 커버.

배경: excel-mcp-server의 multi-step workflow가 Sonnet 4.6의 multi-call 불안정성과
결합되어 반복 실패. 단일 호출 합성 도구 2개(생성/수정)로 실패 지점 원천 차단.

도구 2개:
- create_xlsx(filepath, ...) — 신규 생성 (단일 or 다중 시트)
- modify_xlsx(filepath, operations=[...]) — 기존 파일 수정 (7 ops 배열 기반)

설계 불변식:
1. Single-call completion — 한 번의 호출로 완결 (multi-step 금지)
2. 결정론적 응답 포맷 — ✅ SUCCESS: 접두사 + 파일명/경로 명시
3. Silent rename 금지 — 요청 filepath = 실제 저장 filepath
4. 원자성 — 실패 시 wb.save 하지 않음 (디스크 무변)
"""

import re
from pathlib import Path
from typing import Any, Dict, List, Tuple

from mcp.server import Server
from mcp.server.stdio import stdio_server
from mcp.types import Tool, TextContent

from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, LineChart, PieChart, Reference

OUTPUT_DIR = Path(__file__).parent.parent.parent.parent / "data" / "xlsx_output"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# 시트명 Excel 제약
_INVALID_SHEET_CHARS = set(':\\/?*[]')
_MAX_SHEET_NAME_LEN = 31

# operations 최대 개수 (single-call 원칙 보호)
_MAX_OPERATIONS = 100

# 지원 op 목록 (15종)
_SUPPORTED_OPS = frozenset([
    # 데이터
    "update_cells",
    "apply_formula",
    # 시트
    "add_sheet",
    "delete_sheet",
    "rename_sheet",
    "copy_worksheet",
    # 행·열
    "insert_rows",
    "insert_columns",
    "delete_rows",
    "delete_columns",
    # 서식
    "format_range",
    "merge_cells",
    "unmerge_cells",
    # 차트·피벗
    "create_chart",
    "create_pivot_table",
])

server = Server("xlsx-simple")


# ============================================================================
# 공통 유틸
# ============================================================================
def _resolve_path(filepath: str) -> Path:
    """파일명만 주어지면 OUTPUT_DIR 하위로 해석. 파일명 자체는 불변.

    - 절대 경로 → 그대로 사용
    - 경로 구분자 없는 파일명 → OUTPUT_DIR / filepath
    - 경로 구분자 포함 상대경로 → 그대로 Path (호출자 맥락에 의존)
    - .xlsx 확장자 자동 부여
    """
    p = Path(filepath)
    if not p.is_absolute() and "/" not in filepath and "\\" not in filepath:
        p = OUTPUT_DIR / filepath
    if p.suffix.lower() != ".xlsx":
        p = p.with_suffix(".xlsx")
    return p


def _validate_sheet_name(name: str) -> None:
    """Excel 시트명 제약 검증. 위반 시 ValueError."""
    if not name or not isinstance(name, str):
        raise ValueError(f"시트명은 비어있지 않은 문자열이어야 합니다: {name!r}")
    if len(name) > _MAX_SHEET_NAME_LEN:
        raise ValueError(f"시트명이 31자를 초과합니다: {name!r}")
    bad = [c for c in name if c in _INVALID_SHEET_CHARS]
    if bad:
        raise ValueError(f"시트명에 허용되지 않은 문자 포함: {name!r} (금지: : \\ / ? * [ ])")


def _err(msg: str) -> List[TextContent]:
    return [TextContent(type="text", text=f"Error: {msg}")]


def _ok(msg: str) -> List[TextContent]:
    return [TextContent(type="text", text=msg)]


# ============================================================================
# list_tools
# ============================================================================
@server.list_tools()
async def list_tools() -> List[Tool]:
    return [
        Tool(
            name="create_xlsx",
            description=(
                "엑셀(.xlsx) 파일을 단번에 생성합니다. 새 파일 생성은 **반드시 이 도구 하나만** 호출하세요. "
                "단일 시트: top-level headers/rows 사용. "
                "다중 시트: sheets=[{name, headers, rows}, ...] 배열 사용. "
                "기존 파일 수정에는 modify_xlsx를 쓰세요."
            ),
            inputSchema={
                "type": "object",
                "properties": {
                    "filepath": {
                        "type": "string",
                        "description": "생성할 파일 경로 (.xlsx). 파일명만 주면 xlsx_output/ 하위에 저장됨.",
                    },
                    "headers": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "(단일 시트) 첫 행 헤더 배열. 예: ['A', 'B', 'C', 'D']",
                    },
                    "rows": {
                        "type": "array",
                        "items": {"type": "array"},
                        "description": "(단일 시트) 데이터 행 배열. 예: [[1,2,3,4], [5,6,7,8]]",
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "(단일 시트) 시트명. 기본값 'Sheet'.",
                        "default": "Sheet",
                    },
                    "sheets": {
                        "type": "array",
                        "description": "(다중 시트) 시트 배열. 지정하면 headers/rows 대신 사용됨. 예: [{name:'매출', headers:[...], rows:[[...]]}, {name:'비용', ...}]",
                        "items": {
                            "type": "object",
                            "properties": {
                                "name": {"type": "string"},
                                "headers": {"type": "array", "items": {"type": "string"}},
                                "rows": {"type": "array", "items": {"type": "array"}},
                            },
                            "required": ["name", "headers", "rows"],
                        },
                    },
                },
                "required": ["filepath"],
            },
        ),
        Tool(
            name="modify_xlsx",
            description=(
                "기존 엑셀(.xlsx) 파일을 단번에 수정합니다. 수정은 **반드시 이 도구 하나만** 호출하세요. "
                "여러 변경을 operations 배열로 한 번에 전달하세요. 중간 실패 시 파일은 원본 그대로 유지됩니다(원자성). "
                "지원 op: update_cells, add_sheet, delete_sheet, rename_sheet, apply_formula, delete_rows, delete_columns."
            ),
            inputSchema={
                "type": "object",
                "properties": {
                    "filepath": {
                        "type": "string",
                        "description": "수정할 기존 .xlsx 경로. 파일명만 주면 xlsx_output/ 하위에서 찾음.",
                    },
                    "operations": {
                        "type": "array",
                        "description": (
                            "적용할 작업 배열. 지원 op:\n"
                            "[데이터]\n"
                            "- {op:'update_cells', sheet:'Sheet', start_cell:'A1', values:[['x','y'],[1,2]]}\n"
                            "- {op:'apply_formula', sheet:'Sheet', cell:'C10', formula:'=SUM(C2:C9)'}\n"
                            "[시트]\n"
                            "- {op:'add_sheet', name:'Summary', headers:['A','B'], rows:[[1,2]]}\n"
                            "- {op:'delete_sheet', name:'Sheet2'}\n"
                            "- {op:'rename_sheet', old_name:'Sheet1', new_name:'메인'}\n"
                            "- {op:'copy_worksheet', source:'Sheet1', target:'Sheet1_복사본'}  서식·수식·병합 모두 포함\n"
                            "[행·열]\n"
                            "- {op:'insert_rows', sheet:'Sheet', start_row:2, count:1}\n"
                            "- {op:'insert_columns', sheet:'Sheet', start_col:'B', count:1}\n"
                            "- {op:'delete_rows', sheet:'Sheet', start_row:2, count:3}\n"
                            "- {op:'delete_columns', sheet:'Sheet', start_col:'B', count:1}\n"
                            "[서식]\n"
                            "- {op:'format_range', sheet:'Sheet', start_cell:'A1', end_cell:'D1', bold:true, bg_color:'FFFF00', font_color:'000000', alignment:'center', number_format:'#,##0', border_style:'thin'}\n"
                            "- {op:'merge_cells', sheet:'Sheet', start_cell:'A1', end_cell:'D1'}\n"
                            "- {op:'unmerge_cells', sheet:'Sheet', start_cell:'A1', end_cell:'D1'}\n"
                            "[차트·피벗]\n"
                            "- {op:'create_chart', sheet:'Sheet', chart_type:'bar'|'line'|'pie', data_range:'A1:B10', target_cell:'D2', title:'매출', x_axis:'월', y_axis:'금액'}\n"
                            "- {op:'create_pivot_table', source_sheet:'Data', source_range:'A1:D100', target_sheet:'Pivot', rows:['지역'], values:[['매출','sum']], columns:[], agg_func:'sum'}\n"
                        ),
                        "items": {"type": "object"},
                    },
                },
                "required": ["filepath", "operations"],
            },
        ),
    ]


# ============================================================================
# create_xlsx 핸들러
# ============================================================================
def _handle_create_xlsx(arguments: dict) -> List[TextContent]:
    filepath = arguments.get("filepath", "")
    if not filepath:
        return _err("filepath가 필요합니다.")

    sheets_arg = arguments.get("sheets")

    # 모드 결정: sheets 배열 우선, 없으면 단일 시트 (headers/rows)
    if sheets_arg is not None:
        if not isinstance(sheets_arg, list) or not sheets_arg:
            return _err("sheets는 비어있지 않은 배열이어야 합니다.")
        multi_sheets: List[Dict[str, Any]] = []
        seen_names = set()
        for i, s in enumerate(sheets_arg):
            if not isinstance(s, dict):
                return _err(f"sheets[{i}]는 객체여야 합니다.")
            name = s.get("name", "")
            try:
                _validate_sheet_name(name)
            except ValueError as e:
                return _err(f"sheets[{i}]: {e}")
            if name in seen_names:
                return _err(f"sheets[{i}]: 시트명 중복 '{name}'")
            seen_names.add(name)
            headers = s.get("headers", [])
            rows = s.get("rows", [])
            if not isinstance(headers, list):
                return _err(f"sheets[{i}].headers는 배열이어야 합니다.")
            if not isinstance(rows, list):
                return _err(f"sheets[{i}].rows는 배열이어야 합니다.")
            for r_idx, row in enumerate(rows):
                if not isinstance(row, list):
                    return _err(f"sheets[{i}].rows[{r_idx}]는 배열이어야 합니다.")
            multi_sheets.append({"name": name, "headers": headers, "rows": rows})
    else:
        headers = arguments.get("headers", [])
        rows = arguments.get("rows", [])
        sheet_name = arguments.get("sheet_name", "Sheet")
        if not isinstance(headers, list) or not headers:
            return _err("headers는 비어있지 않은 배열이어야 합니다 (또는 sheets 배열 사용).")
        if not isinstance(rows, list):
            return _err("rows는 배열이어야 합니다.")
        for r_idx, row in enumerate(rows):
            if not isinstance(row, list):
                return _err(f"rows[{r_idx}]는 배열이어야 합니다.")
        try:
            _validate_sheet_name(sheet_name)
        except ValueError as e:
            return _err(str(e))
        multi_sheets = [{"name": sheet_name, "headers": headers, "rows": rows}]

    p = _resolve_path(filepath)

    try:
        wb = Workbook()
        # 첫 시트는 기본 active 시트 재사용
        first = multi_sheets[0]
        ws = wb.active
        ws.title = first["name"]
        ws.append([str(h) for h in first["headers"]])
        for row in first["rows"]:
            ws.append(row)
        # 이후 시트는 create_sheet
        for s in multi_sheets[1:]:
            ws = wb.create_sheet(s["name"])
            ws.append([str(h) for h in s["headers"]])
            for row in s["rows"]:
                ws.append(row)
        p.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(p))
    except Exception as e:
        return _err(f"엑셀 생성 실패 - {type(e).__name__}: {e}")

    # 응답
    if len(multi_sheets) == 1:
        s = multi_sheets[0]
        size_txt = f"{len(s['rows']) + 1}행 × {len(s['headers'])}열 (헤더 포함)"
        msg = (
            f"✅ SUCCESS: 엑셀 파일 생성 완료\n"
            f"- 파일명: {p.name}\n"
            f"- 경로: {str(p).replace(chr(92), '/')}\n"
            f"- 시트: '{s['name']}'\n"
            f"- 작성 범위: {size_txt}\n"
            f"작업 완료. 사용자에게 `**파일명:** {p.name}` 형태로 안내하세요."
        )
    else:
        lines = [
            f"✅ SUCCESS: 엑셀 파일 생성 완료",
            f"- 파일명: {p.name}",
            f"- 경로: {str(p).replace(chr(92), '/')}",
            f"- 시트 수: {len(multi_sheets)}",
            f"- 시트별 작성 범위:",
        ]
        for s in multi_sheets:
            lines.append(
                f"    * '{s['name']}': {len(s['rows']) + 1}행 × {len(s['headers'])}열 (헤더 포함)"
            )
        lines.append(f"작업 완료. 사용자에게 `**파일명:** {p.name}` 형태로 안내하세요.")
        msg = "\n".join(lines)
    return _ok(msg)


# ============================================================================
# modify_xlsx 핸들러
# ============================================================================
def _apply_op(wb: Workbook, op: Dict[str, Any]) -> str:
    """단일 op를 wb에 적용. 성공 시 사람이 읽을 수 있는 요약 문자열 반환. 실패 시 예외."""
    op_type = op.get("op")
    if op_type not in _SUPPORTED_OPS:
        raise ValueError(f"지원하지 않는 op: {op_type!r} (지원: {sorted(_SUPPORTED_OPS)})")

    if op_type == "update_cells":
        sheet = op.get("sheet", "")
        start_cell = op.get("start_cell", "")
        values = op.get("values", [])
        if sheet not in wb.sheetnames:
            raise ValueError(f"시트 없음: {sheet!r}")
        if not isinstance(values, list) or not values:
            raise ValueError("values는 비어있지 않은 2차원 배열이어야 합니다.")
        ws = wb[sheet]
        # start_cell 좌표 계산
        anchor = ws[start_cell]
        base_row, base_col = anchor.row, anchor.column
        for r_off, row in enumerate(values):
            if not isinstance(row, list):
                raise ValueError(f"values[{r_off}]는 배열이어야 합니다.")
            for c_off, v in enumerate(row):
                ws.cell(row=base_row + r_off, column=base_col + c_off, value=v)
        return f"update_cells @ '{sheet}'!{start_cell} ({len(values)}행)"

    if op_type == "add_sheet":
        name = op.get("name", "")
        _validate_sheet_name(name)
        if name in wb.sheetnames:
            raise ValueError(f"시트 '{name}' 이미 존재")
        headers = op.get("headers", [])
        rows = op.get("rows", [])
        if not isinstance(headers, list):
            raise ValueError("headers는 배열이어야 합니다.")
        if not isinstance(rows, list):
            raise ValueError("rows는 배열이어야 합니다.")
        ws = wb.create_sheet(name)
        if headers:
            ws.append([str(h) for h in headers])
        for r_idx, row in enumerate(rows):
            if not isinstance(row, list):
                raise ValueError(f"rows[{r_idx}]는 배열이어야 합니다.")
            ws.append(row)
        return f"add_sheet '{name}' ({len(rows)}행)"

    if op_type == "delete_sheet":
        name = op.get("name", "")
        if name not in wb.sheetnames:
            raise ValueError(f"시트 없음: {name!r}")
        if len(wb.sheetnames) == 1:
            raise ValueError("마지막 시트는 삭제할 수 없습니다.")
        del wb[name]
        return f"delete_sheet '{name}'"

    if op_type == "rename_sheet":
        old_name = op.get("old_name", "")
        new_name = op.get("new_name", "")
        if old_name not in wb.sheetnames:
            raise ValueError(f"시트 없음: {old_name!r}")
        _validate_sheet_name(new_name)
        if new_name in wb.sheetnames and new_name != old_name:
            raise ValueError(f"시트 '{new_name}' 이미 존재")
        wb[old_name].title = new_name
        return f"rename_sheet '{old_name}' → '{new_name}'"

    if op_type == "apply_formula":
        sheet = op.get("sheet", "")
        cell = op.get("cell", "")
        formula = op.get("formula", "")
        if sheet not in wb.sheetnames:
            raise ValueError(f"시트 없음: {sheet!r}")
        if not isinstance(formula, str) or not formula.strip():
            raise ValueError("formula는 비어있지 않은 문자열이어야 합니다.")
        if not formula.startswith("="):
            formula = "=" + formula
        wb[sheet][cell] = formula
        return f"apply_formula @ '{sheet}'!{cell}"

    if op_type == "delete_rows":
        sheet = op.get("sheet", "")
        start_row = op.get("start_row")
        count = op.get("count", 1)
        if sheet not in wb.sheetnames:
            raise ValueError(f"시트 없음: {sheet!r}")
        if not isinstance(start_row, int) or start_row < 1:
            raise ValueError(f"start_row는 1 이상의 정수여야 합니다: {start_row!r}")
        if not isinstance(count, int) or count < 1:
            raise ValueError(f"count는 1 이상의 정수여야 합니다: {count!r}")
        wb[sheet].delete_rows(start_row, count)
        return f"delete_rows @ '{sheet}' {start_row}부터 {count}개"

    if op_type == "delete_columns":
        sheet = op.get("sheet", "")
        start_col = op.get("start_col")
        count = op.get("count", 1)
        if sheet not in wb.sheetnames:
            raise ValueError(f"시트 없음: {sheet!r}")
        col_idx = _resolve_column(start_col)
        if not isinstance(count, int) or count < 1:
            raise ValueError(f"count는 1 이상의 정수여야 합니다: {count!r}")
        wb[sheet].delete_cols(col_idx, count)
        return f"delete_columns @ '{sheet}' {start_col}부터 {count}개"

    if op_type == "copy_worksheet":
        source = op.get("source", "")
        target = op.get("target", "")
        if source not in wb.sheetnames:
            raise ValueError(f"원본 시트 없음: {source!r}")
        _validate_sheet_name(target)
        if target in wb.sheetnames:
            raise ValueError(f"대상 시트 '{target}' 이미 존재")
        new_ws = wb.copy_worksheet(wb[source])
        new_ws.title = target
        return f"copy_worksheet '{source}' → '{target}' (서식·수식 포함)"

    if op_type == "insert_rows":
        sheet = op.get("sheet", "")
        start_row = op.get("start_row")
        count = op.get("count", 1)
        if sheet not in wb.sheetnames:
            raise ValueError(f"시트 없음: {sheet!r}")
        if not isinstance(start_row, int) or start_row < 1:
            raise ValueError(f"start_row는 1 이상의 정수여야 합니다: {start_row!r}")
        if not isinstance(count, int) or count < 1:
            raise ValueError(f"count는 1 이상의 정수여야 합니다: {count!r}")
        wb[sheet].insert_rows(start_row, count)
        return f"insert_rows @ '{sheet}' {start_row}위치 {count}개"

    if op_type == "insert_columns":
        sheet = op.get("sheet", "")
        start_col = op.get("start_col")
        count = op.get("count", 1)
        if sheet not in wb.sheetnames:
            raise ValueError(f"시트 없음: {sheet!r}")
        col_idx = _resolve_column(start_col)
        if not isinstance(count, int) or count < 1:
            raise ValueError(f"count는 1 이상의 정수여야 합니다: {count!r}")
        wb[sheet].insert_cols(col_idx, count)
        return f"insert_columns @ '{sheet}' {start_col}위치 {count}개"

    if op_type == "format_range":
        sheet = op.get("sheet", "")
        start_cell = op.get("start_cell", "")
        end_cell = op.get("end_cell") or start_cell
        if sheet not in wb.sheetnames:
            raise ValueError(f"시트 없음: {sheet!r}")
        ws = wb[sheet]
        cell_range = f"{start_cell}:{end_cell}"
        font_kwargs = {}
        if "bold" in op:
            font_kwargs["bold"] = bool(op["bold"])
        if "italic" in op:
            font_kwargs["italic"] = bool(op["italic"])
        if "underline" in op and op["underline"]:
            font_kwargs["underline"] = "single"
        if "font_size" in op:
            font_kwargs["size"] = int(op["font_size"])
        if "font_color" in op and op["font_color"]:
            font_kwargs["color"] = _normalize_color(op["font_color"])
        fill = None
        if "bg_color" in op and op["bg_color"]:
            fill = PatternFill(start_color=_normalize_color(op["bg_color"]),
                               end_color=_normalize_color(op["bg_color"]),
                               fill_type="solid")
        border = None
        if "border_style" in op and op["border_style"]:
            side = Side(style=op["border_style"],
                        color=_normalize_color(op.get("border_color", "000000")))
            border = Border(left=side, right=side, top=side, bottom=side)
        align = None
        if "alignment" in op and op["alignment"]:
            align = Alignment(horizontal=op["alignment"],
                              wrap_text=bool(op.get("wrap_text", False)))
        elif "wrap_text" in op:
            align = Alignment(wrap_text=bool(op["wrap_text"]))
        number_format = op.get("number_format")
        for row in ws[cell_range]:
            for cell in row:
                if font_kwargs:
                    cell.font = Font(**font_kwargs)
                if fill is not None:
                    cell.fill = fill
                if border is not None:
                    cell.border = border
                if align is not None:
                    cell.alignment = align
                if number_format:
                    cell.number_format = number_format
        return f"format_range @ '{sheet}'!{cell_range}"

    if op_type == "merge_cells":
        sheet = op.get("sheet", "")
        start_cell = op.get("start_cell", "")
        end_cell = op.get("end_cell", "")
        if sheet not in wb.sheetnames:
            raise ValueError(f"시트 없음: {sheet!r}")
        if not start_cell or not end_cell:
            raise ValueError("start_cell, end_cell 모두 필요합니다.")
        wb[sheet].merge_cells(f"{start_cell}:{end_cell}")
        return f"merge_cells @ '{sheet}'!{start_cell}:{end_cell}"

    if op_type == "unmerge_cells":
        sheet = op.get("sheet", "")
        start_cell = op.get("start_cell", "")
        end_cell = op.get("end_cell", "")
        if sheet not in wb.sheetnames:
            raise ValueError(f"시트 없음: {sheet!r}")
        if not start_cell or not end_cell:
            raise ValueError("start_cell, end_cell 모두 필요합니다.")
        wb[sheet].unmerge_cells(f"{start_cell}:{end_cell}")
        return f"unmerge_cells @ '{sheet}'!{start_cell}:{end_cell}"

    if op_type == "create_chart":
        sheet = op.get("sheet", "")
        chart_type = (op.get("chart_type", "") or "").lower()
        data_range = op.get("data_range", "")
        target_cell = op.get("target_cell", "")
        title = op.get("title", "")
        x_axis = op.get("x_axis", "")
        y_axis = op.get("y_axis", "")
        if sheet not in wb.sheetnames:
            raise ValueError(f"시트 없음: {sheet!r}")
        chart_map = {"bar": BarChart, "line": LineChart, "pie": PieChart}
        if chart_type not in chart_map:
            raise ValueError(f"지원하지 않는 chart_type: {chart_type!r} (지원: bar, line, pie)")
        if not data_range or ":" not in data_range:
            raise ValueError(f"data_range는 'A1:B10' 형식이어야 합니다: {data_range!r}")
        if not target_cell:
            raise ValueError("target_cell이 필요합니다 (차트 위치).")
        ws = wb[sheet]
        chart = chart_map[chart_type]()
        if title:
            chart.title = title
        if chart_type != "pie":
            if x_axis:
                chart.x_axis.title = x_axis
            if y_axis:
                chart.y_axis.title = y_axis
        # data_range를 Reference로 변환. 첫 행을 헤더/카테고리로 자동 인식.
        ref = Reference(ws, range_string=f"{sheet}!{data_range}")
        chart.add_data(ref, titles_from_data=True)
        ws.add_chart(chart, target_cell)
        return f"create_chart({chart_type}) @ '{sheet}'!{target_cell} from {data_range}"

    if op_type == "create_pivot_table":
        # openpyxl은 진짜 Excel PivotTable 객체 생성이 제한적이므로,
        # pandas로 집계 계산 후 결과를 일반 시트로 삽입 (값+헤더 형태).
        source_sheet = op.get("source_sheet", "")
        source_range = op.get("source_range", "")
        target_sheet = op.get("target_sheet", "")
        rows_cfg = op.get("rows", [])
        values_cfg = op.get("values", [])
        columns_cfg = op.get("columns", []) or []
        agg_func_default = op.get("agg_func", "sum")
        if source_sheet not in wb.sheetnames:
            raise ValueError(f"원본 시트 없음: {source_sheet!r}")
        _validate_sheet_name(target_sheet)
        if target_sheet in wb.sheetnames:
            raise ValueError(f"대상 시트 '{target_sheet}' 이미 존재")
        if not source_range or ":" not in source_range:
            raise ValueError(f"source_range는 'A1:D100' 형식이어야 합니다: {source_range!r}")
        if not rows_cfg or not values_cfg:
            raise ValueError("rows와 values는 비어있을 수 없습니다.")
        try:
            import pandas as pd
        except ImportError:
            raise RuntimeError("pandas가 필요합니다 (create_pivot_table)")
        src_ws = wb[source_sheet]
        raw = list(src_ws[source_range])
        if len(raw) < 2:
            raise ValueError("source_range에 헤더 + 1행 이상의 데이터가 필요합니다.")
        headers = [c.value for c in raw[0]]
        data_rows = [[c.value for c in r] for r in raw[1:]]
        df = pd.DataFrame(data_rows, columns=headers)
        # values_cfg: [["col", "agg"]] or ["col"] → agg_func_default
        vals: List[str] = []
        aggs: Dict[str, Any] = {}
        for v in values_cfg:
            if isinstance(v, list) and len(v) >= 2:
                col, agg = v[0], v[1]
            elif isinstance(v, str):
                col, agg = v, agg_func_default
            else:
                raise ValueError(f"values 항목 형식 오류: {v!r}")
            vals.append(col)
            aggs[col] = agg
        pivot = pd.pivot_table(
            df,
            index=rows_cfg,
            columns=columns_cfg if columns_cfg else None,
            values=vals,
            aggfunc=aggs,
            fill_value=0,
        )
        # MultiIndex columns를 평탄화
        if hasattr(pivot.columns, "to_flat_index"):
            flat_cols = [" / ".join(str(x) for x in c) if isinstance(c, tuple) else str(c)
                         for c in pivot.columns.to_flat_index()]
        else:
            flat_cols = [str(c) for c in pivot.columns]
        pivot_reset = pivot.reset_index()
        new_ws = wb.create_sheet(target_sheet)
        new_ws.append(list(rows_cfg) + flat_cols)
        for row in pivot_reset.itertuples(index=False):
            new_ws.append(list(row))
        return f"create_pivot_table '{source_sheet}' → '{target_sheet}' ({len(pivot_reset)}행)"

    # 도달 불가 (상단에서 이미 검증)
    raise ValueError(f"내부 오류: op 처리 누락 {op_type!r}")


def _resolve_column(val) -> int:
    """start_col: 문자('A', 'B') 또는 정수(1, 2) → 1-based int 인덱스."""
    if isinstance(val, str):
        try:
            return column_index_from_string(val.upper())
        except Exception:
            raise ValueError(f"start_col 문자열이 유효하지 않습니다: {val!r}")
    if isinstance(val, int) and val >= 1:
        return val
    raise ValueError(f"start_col은 'A' 같은 문자 또는 1 이상의 정수여야 합니다: {val!r}")


def _normalize_color(color: str) -> str:
    """#RRGGBB, RRGGBB, RGB → openpyxl의 8자리 AARRGGBB (alpha FF 고정)."""
    if not isinstance(color, str):
        raise ValueError(f"color는 문자열이어야 합니다: {color!r}")
    s = color.strip().lstrip("#").upper()
    if len(s) == 6:
        return "FF" + s
    if len(s) == 8:
        return s
    if len(s) == 3:
        return "FF" + "".join(ch * 2 for ch in s)
    raise ValueError(f"color 포맷이 올바르지 않습니다 (예: '#FFFF00'): {color!r}")


def _handle_modify_xlsx(arguments: dict) -> List[TextContent]:
    filepath = arguments.get("filepath", "")
    operations = arguments.get("operations", [])

    if not filepath:
        return _err("filepath가 필요합니다.")
    if not isinstance(operations, list) or not operations:
        return _err("operations는 비어있지 않은 배열이어야 합니다.")
    if len(operations) > _MAX_OPERATIONS:
        return _err(f"operations 최대 {_MAX_OPERATIONS}개 (현재 {len(operations)}개)")

    p = _resolve_path(filepath)
    if not p.exists():
        return _err(f"파일이 존재하지 않습니다: {str(p).replace(chr(92), '/')}")

    try:
        wb = load_workbook(str(p))
    except Exception as e:
        return _err(f"파일 로드 실패 - {type(e).__name__}: {e}")

    applied: List[str] = []
    for idx, op in enumerate(operations):
        if not isinstance(op, dict):
            return _err(f"operations[{idx}]는 객체여야 합니다.")
        try:
            summary = _apply_op(wb, op)
        except Exception as e:
            return _err(
                f"operations[{idx}] op='{op.get('op', '?')}' 실패 - {type(e).__name__}: {e}"
            )
        applied.append(summary)

    try:
        wb.save(str(p))
    except Exception as e:
        return _err(f"저장 실패 - {type(e).__name__}: {e}")

    size = p.stat().st_size
    lines = [
        f"✅ SUCCESS: 엑셀 파일 수정 완료",
        f"- 파일명: {p.name}",
        f"- 경로: {str(p).replace(chr(92), '/')}",
        f"- 적용된 작업: {len(applied)}개",
    ]
    for i, s in enumerate(applied, 1):
        lines.append(f"    {i}. {s}")
    lines.append(f"- 파일 크기: {size:,} bytes")
    lines.append(f"작업 완료. 사용자에게 `**파일명:** {p.name}` 형태로 안내하세요.")
    return _ok("\n".join(lines))


# ============================================================================
# call_tool dispatcher
# ============================================================================
@server.call_tool()
async def call_tool(name: str, arguments: dict) -> List[TextContent]:
    if name == "create_xlsx":
        return _handle_create_xlsx(arguments)
    if name == "modify_xlsx":
        return _handle_modify_xlsx(arguments)
    return _err(f"Unknown tool '{name}'")


# ============================================================================
# main
# ============================================================================
async def main():
    async with stdio_server() as (read_stream, write_stream):
        await server.run(read_stream, write_stream, server.create_initialization_options())


if __name__ == "__main__":
    import asyncio
    asyncio.run(main())
